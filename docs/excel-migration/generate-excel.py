"""Generate a realistic legacy NGO case Excel file for Kaval migration."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Cases"

# ── Column headers ──
headers = [
    "Case ID",          # A
    "CR Number",        # B  -> crimeNumber
    "ST Number",        # C  -> stNumber
    "Beneficiary Name", # D  -> beneficiaryName
    "Age",              # E  -> beneficiaryAge
    "Contact No",       # F  -> beneficiaryContact
    "Address",          # G  -> LEGACY ONLY
    "Type of Offence",  # H  -> typeOfOffence
    "Classification",   # I  -> offenceClassification
    "Domicile/Area",    # J  -> domicile
    "First Offender",   # K  -> isFirstTimeOffender
    "Stage",            # L  -> LEGACY ONLY
    "Assigned Worker",  # M  -> LEGACY ONLY
    "Date Registered",  # N  -> LEGACY ONLY
    "Remarks",          # O  -> LEGACY ONLY
]

header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="0D6E6E", end_color="0D6E6E", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# ── Data rows (realistic juvenile justice NGO cases) ──
data = [
    # (CaseID, CR, ST, Name, Age, Contact, Address, Offence, Class, Domicile, FirstOff, Stage, Worker, DateReg, Remarks)
    ("KAV-2024-001", "CR-2024-001234", "ST-2024-5678",  "Ravi Kumar",       17, "9876543210", "123 MG Road, Bangalore",          "Theft",                    "Petty",   "Urban",  "Yes", "Initial Assessment",  "Anita Sharma",   "2024-01-15", ""),
    ("KAV-2024-002", "CR-2024-001235", "ST-2024-5679",  "Sunita Devi",       15, "9876543211", "Village P.O. Barabanki, UP",      "Pickpocketing",            "Petty",   "Rural",  "Yes", "Social Investigation","Rajesh Patel",   "2024-01-20", "First-time offender, family cooperative"),
    ("KAV-2024-003", "CR-2024-001236", "ST-2024-5680",  "Mohammad Salim",    16, "9876543212", "56 Zakir Nagar, Delhi",            "Chain Snatching",          "Serious", "Urban",  "No",  "Counseling",        "Fatima Khan",    "2024-02-01", "Repeat offender — previous case closed"),
    ("KAV-2024-004", "CR-2024-001237", "ST-2024-5681",  "Lakshmi Priya",     14, "9876543213", "Coastal Village, Puri, Odisha",    "Theft",                    "Petty",   "Coastal","Yes", "Probation",        "Prakash Rao",    "2024-02-10", ""),
    ("KAV-2024-005", "CR-2024-001238", "ST-2024-5682",  "Arjun Singh",       16, "9876543214", "Tribal Hamlet, Jhabua, MP",        "Assault",                  "Grave",   "Tribal", "Yes", "Institutional Care", "Meena Verma",    "2024-03-05", "Tribal welfare board involved"),
    ("KAV-2024-006", "CR-2024-001239", "ST-2024-5683",  "Priya Sharma",       12, "",            "Dharavi Slum, Mumbai",             "Begging",                  "Petty",   "Slum",   "Yes", "Rehabilitation",   "Suresh Iyer",    "2024-03-12", "Child in need of care — POCSO flagged"),
    ("KAV-2024-007", "CR-2024-001240", "ST-2024-5684",  "Deepak Yadav",       17, "9876543215", "456 Civil Lines, Lucknow",         "Robbery",                  "Serious", "Urban",  "No",  "Judicial Process",  "Anita Sharma",   "2024-03-20", ""),
    ("KAV-2024-008", "CR-2024-001241", "ST-2024-5685",  "Geeta Kumari",       15, "9876543216", "Main Bazaar, Patna",              "Theft",                    "Minor",   "Urban",  "Yes", "Social Investigation","Rajesh Patel",  "2024-04-01", ""),
    ("KAV-2024-009", "CR-2024-001242", "ST-2024-5686",  "Suresh Naik",        16, "9876543217", "Hill Colony, Ooty, Tamil Nadu",    "Criminal Trespass",        "Grave",   "Rural",  "No",  "Counseling",        "Prakash Rao",    "2024-04-15", "Land dispute related"),
    ("KAV-2024-010", "CR-2024-001243", "ST-2024-5687",  "Aisha Begum",        14, "9876543218", "22 Old City, Hyderabad",           "Pickpocketing",            "Petty",   "Urban",  "Yes", "Probation",        "Fatima Khan",    "2024-05-01", ""),
    ("KAV-2024-011", "CR-2024-001244", "ST-2024-5688",  "Mohan Das",          17, "",            "Riverbank Settlement, Guwahati",   "Armed Robbery",             "Heinous/Grave","Coastal","No", "Institutional Care","Meena Verma",  "2024-05-10", "Weapon involved — high risk"),
    ("KAV-2024-012", "CR-2024-001245", "ST-2024-5689",  "Rekha Devi",         13, "9876543219", "Slum Colony, Chennai",             "Begging",                  "Petty",   "Slum",   "Yes", "Rehabilitation",   "Suresh Iyer",    "2024-05-20", "Child labour rescue case"),
    ("KAV-2024-013", "CR-2024-001246", "ST-2024-5690",  "Vijay Kumar",        16, "9876543220", "Mountain Village, Shimla",         "Assault",                  "Grave",   "Rural",  "No",  "Judicial Process",  "Anita Sharma",   "2024-06-01", ""),
    ("KAV-2024-014", "CR-2024-001247", "ST-2024-5691",  "Sita Devi",          14, "9876543221", "Coastal Village, Kanyakumari",     "Theft of Livestock",       "Petty",   "Coastal","Yes", "Initial Assessment", "Rajesh Patel",  "2024-06-10", ""),
    ("KAV-2024-015", "CR-2024-001248", "ST-2024-5692",  "Amit Jha",           17, "9876543222", "Tribal Settlement, Ranchi",        "Hurt/Grievous Hurt",       "Serious", "Tribal", "Yes", "Social Investigation","Prakash Rao",  "2024-06-15", "First offence, tribal council mediation"),
    ("KAV-2024-016", "CR-2024-001249", "ST-2024-5693",  "Kavita Sharma",      "",  "9876543223", "Sector 12, Noida",                "Cyber Crime",              "Serious", "Urban",  "Yes", "Counseling",        "Fatima Khan",    "2024-07-01", "Online fraud — digital literacy case"),
    ("KAV-2024-017", "CR-2024-001250", "ST-2024-5694",  "Bhola Ram",          15,  "",            "Desert Village, Jaisalmer",        "Theft of Cattle",          "Minor",   "Rural",  "Yes", "Probation",        "Suresh Iyer",    "2024-07-15", ""),
]

data_font = Font(name="Calibri", size=10)
data_align = Alignment(vertical="center", wrap_text=True)

for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = data_font
        cell.alignment = data_align
        cell.border = thin_border

# ── Column widths ──
col_widths = [16, 18, 18, 22, 6, 14, 28, 22, 16, 16, 14, 20, 18, 16, 40]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# ── Freeze top row ──
ws.freeze_panes = "A2"

# ── Auto-filter ──
ws.auto_filter.ref = f"A1:O{len(data) + 1}"

# ── Legend / Notes sheet ──
ws2 = wb.create_sheet("Legend")
ws2.cell(row=1, column=1, value="Field Mapping Reference").font = Font(bold=True, size=14)
ws2.cell(row=3, column=1, value="Legacy Column").font = Font(bold=True)
ws2.cell(row=3, column=2, value="Maps To (Kaval Field)").font = Font(bold=True)
ws2.cell(row=3, column=3, value="Status").font = Font(bold=True)

legend_data = [
    ("CR Number", "crimeNumber", "IMPORTED"),
    ("ST Number", "stNumber", "IMPORTED"),
    ("Beneficiary Name", "beneficiaryName", "IMPORTED"),
    ("Age", "beneficiaryAge", "IMPORTED (optional)"),
    ("Contact No", "beneficiaryContact", "IMPORTED (optional)"),
    ("Type of Offence", "typeOfOffence", "IMPORTED"),
    ("Classification", "offenceClassification", "IMPORTED (enum mapped)"),
    ("Domicile/Area", "domicile", "IMPORTED (enum mapped)"),
    ("First Offender", "isFirstTimeOffender", "IMPORTED (optional)"),
    ("Case ID", "—", "LEGACY ONLY"),
    ("Address", "—", "LEGACY ONLY"),
    ("Stage", "—", "LEGACY ONLY"),
    ("Assigned Worker", "—", "LEGACY ONLY"),
    ("Date Registered", "—", "LEGACY ONLY"),
    ("Remarks", "—", "LEGACY ONLY"),
]

green_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
amber_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
grey_fill  = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

for i, (legacy, target, status) in enumerate(legend_data, 4):
    ws2.cell(row=i, column=1, value=legacy).border = thin_border
    ws2.cell(row=i, column=2, value=target).border = thin_border
    c = ws2.cell(row=i, column=3, value=status)
    c.border = thin_border
    if status == "IMPORTED":
        c.fill = green_fill
    elif "(optional)" in status or "enum" in status:
        c.fill = amber_fill
    else:
        c.fill = grey_fill

ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 30
ws2.column_dimensions["C"].width = 22

# ── Save ──
filepath = "docs/excel-migration/legacy-cases-export.xlsx"
wb.save(filepath)
print(f"[OK] Created {filepath} with {len(data)} realistic case rows")
print(f"  Sheets: {wb.sheetnames}")
